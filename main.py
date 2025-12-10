import os
import sys
import traceback
from datetime import datetime

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from PIL import Image, ImageTk
from tkinterdnd2 import DND_FILES, TkinterDnD
from openpyxl import Workbook, load_workbook

BREEDS_FILE = "breeds.txt"
EXCEL_FILE = "customer_data.xlsx"
OUTPUT_ROOT = "고객사진"


# ---------- 공통 유틸 ----------

def load_breeds_or_die(filename: str = BREEDS_FILE):
    """
    품종 리스트를 읽어오고, 실패하면 오류창을 띄우고 프로그램을 종료한다.
    """
    try:
        breeds = []
        with open(filename, "r", encoding="utf-8") as f:
            for line in f:
                name = line.strip()
                if name:
                    breeds.append(name)

        # '기타(직접입력)' 없으면 자동 추가
        if "기타(직접입력)" not in breeds:
            breeds.append("기타(직접입력)")

        return breeds

    except Exception:
        err_text = traceback.format_exc()

        root = tk.Tk()
        root.title("치명적 오류")
        root.geometry("520x260")
        root.resizable(False, False)

        frame = ttk.Frame(root, padding=10)
        frame.pack(fill="both", expand=True)

        msg = (
            f"품종 리스트 파일을 불러오는 중 오류가 발생했습니다.\n"
            f"파일 이름: {filename}\n\n"
            "프로그램을 종료합니다."
        )
        label = ttk.Label(frame, text=msg, justify="left")
        label.pack(anchor="w")

        text_box = tk.Text(frame, height=6, wrap="word")
        text_box.pack(fill="both", expand=True, pady=5)
        text_box.insert("1.0", err_text)
        text_box.config(state="disabled")

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(anchor="e", pady=(5, 0))

        def copy_error():
            root.clipboard_clear()
            root.clipboard_append(err_text)
            messagebox.showinfo("복사 완료", "오류 기록이 클립보드에 복사되었습니다.", parent=root)

        def close_app():
            root.destroy()
            sys.exit(1)

        ttk.Button(btn_frame, text="오류 기록 복사", command=copy_error).pack(side="right", padx=5)
        ttk.Button(btn_frame, text="확인", command=close_app).pack(side="right")

        root.mainloop()
        # 여기까지 오면 이미 종료됨
        sys.exit(1)


def sanitize_for_path(name: str) -> str:
    """폴더/파일명에 사용할 문자열에서 금지 문자를 제거."""
    invalid = '\\/:*?"<>|'
    result = "".join("_" if ch in invalid else ch for ch in name)
    return result.strip()


def ensure_excel_file(path: str):
    """엑셀 파일이 없으면 새로 만들고, 있으면 그대로 둔다."""
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "고객기록"
        ws.append(
            [
                "기록시각",
                "고객번호",
                "보호자 이름",
                "강아지 이름",
                "품종",
                "오늘 미용 스타일",
                "고객 요구사항",
                "미용 중 특이사항",
                "애프터 케어",
                "결제금액",
                "결제상태",
                "미용 전 사진파일명",
                "미용 후 사진파일명",
            ]
        )
        wb.save(path)


def append_excel_row(path: str, row):
    ensure_excel_file(path)
    wb = load_workbook(path)
    ws = wb.active
    ws.append(row)
    wb.save(path)


def load_image_for_thumbnail(path, size=(320, 220)):
    """PNG 투명도 포함 이미지를 썸네일용 ImageTk.PhotoImage로 변환."""
    im = Image.open(path)
    if im.mode in ("RGBA", "LA"):
        bg = Image.new("RGB", im.size, (255, 255, 255))
        alpha = im.split()[-1]
        bg.paste(im, mask=alpha)
        im = bg
    else:
        im = im.convert("RGB")

    thumb = im.copy()
    thumb.thumbnail(size)
    return ImageTk.PhotoImage(thumb)


# ---------- 메인 앱 ----------

class DogPhotoTool(TkinterDnD.Tk):
    def __init__(self, breeds):
        super().__init__()

        self.title("강아지 미용 사진 정리 도우미")
        self.geometry("820x640")
        self.resizable(False, False)

        # 상태 변수들
        self.before_path = tk.StringVar()
        self.after_path = tk.StringVar()

        self.dog_name = tk.StringVar()
        self.owner_name = tk.StringVar()
        self.customer_raw = tk.StringVar()
        self.style_today = tk.StringVar()

        self.breed_var = tk.StringVar()
        self.breed_other_var = tk.StringVar()

        self.payment_display = tk.StringVar()
        self.payment_state = tk.StringVar(value="paid")  # "paid" / "pending"

        self.before_thumb = None
        self.after_thumb = None

        self.breeds = breeds

        self._build_ui()
        self._lock_window_size()

    # ---------- UI 구성 ----------

    def _build_ui(self):
        # 전체 패딩 프레임
        root_frame = ttk.Frame(self, padding=10)
        root_frame.pack(fill="both", expand=True)

        # UI를 영역별로 분리하여 유지보수성을 높임
        self._build_photo_area(root_frame)
        self._build_customer_and_pet_area(root_frame)
        self._build_notes_area(root_frame)
        self._build_action_area(root_frame)

        # 초기 품종 상태 반영
        self.on_breed_selected()

    def _build_photo_area(self, root_frame):
        photo_frame = ttk.LabelFrame(root_frame, text="사진 드래그 앤 드롭")
        photo_frame.pack(fill="x", padx=5, pady=5)

        # 고정 크기 컨테이너로 이미지 삽입 시 레이아웃이 변하지 않도록 함
        before_container = ttk.Frame(photo_frame, width=360, height=240)
        after_container = ttk.Frame(photo_frame, width=360, height=240)
        before_container.pack(side="left", padx=5, pady=5)
        after_container.pack(side="left", padx=5, pady=5)
        before_container.pack_propagate(False)
        after_container.pack_propagate(False)

        self.before_label = ttk.Label(
            before_container,
            text="미용 전 사진\n(여기로 드롭 또는 클릭)",
            relief="sunken",
            anchor="center",
        )
        self.before_label.pack(fill="both", expand=True)

        self.after_label = ttk.Label(
            after_container,
            text="미용 후 사진\n(여기로 드롭 또는 클릭)",
            relief="sunken",
            anchor="center",
        )
        self.after_label.pack(fill="both", expand=True)

        # DnD 등록
        self.before_label.drop_target_register(DND_FILES)
        self.after_label.drop_target_register(DND_FILES)
        self.before_label.dnd_bind("<<Drop>>", lambda e: self.on_photo_drop(e, "before"))
        self.after_label.dnd_bind("<<Drop>>", lambda e: self.on_photo_drop(e, "after"))

        # 클릭으로 파일 선택
        self.before_label.bind("<Button-1>", lambda e: self.on_photo_click("before"))
        self.after_label.bind("<Button-1>", lambda e: self.on_photo_click("after"))

    def _build_customer_and_pet_area(self, root_frame):
        info_frame = ttk.LabelFrame(root_frame, text="고객 / 강아지 정보")
        info_frame.pack(fill="x", padx=5, pady=5)

        # 1행: 강아지/보호자 이름
        row1 = ttk.Frame(info_frame)
        row1.pack(fill="x", pady=2)

        ttk.Label(row1, text="강아지 이름", width=14, anchor="e").pack(side="left", padx=(0, 3))
        ttk.Entry(row1, textvariable=self.dog_name, width=20).pack(side="left")

        ttk.Label(row1, text="보호자 이름", width=14, anchor="e").pack(side="left", padx=(20, 3))
        ttk.Entry(row1, textvariable=self.owner_name, width=20).pack(side="left")

        # 2행: 고객번호 / 오늘 미용 스타일
        row2 = ttk.Frame(info_frame)
        row2.pack(fill="x", pady=2)

        ttk.Label(row2, text="고객번호(전화번호)", width=14, anchor="e").pack(side="left", padx=(0, 3))
        self.customer_entry = ttk.Entry(row2, textvariable=self.customer_raw, width=20)
        self.customer_entry.pack(side="left")
        self.customer_entry.insert(0, "010-0000-0000")
        self.customer_entry.config(foreground="gray")
        self.customer_entry.bind("<FocusIn>", self._on_customer_focus_in)
        self.customer_entry.bind("<FocusOut>", self._on_customer_focus_out)

        ttk.Label(row2, text="오늘 미용 스타일", width=14, anchor="e").pack(side="left", padx=(20, 3))
        ttk.Entry(row2, textvariable=self.style_today, width=20).pack(side="left")

        # 3행: 품종 / 결제금액+상태
        row3 = ttk.Frame(info_frame)
        row3.pack(fill="x", pady=2)

        ttk.Label(row3, text="품종", width=14, anchor="e").pack(side="left", padx=(0, 3))

        self.breed_combo = ttk.Combobox(
            row3,
            textvariable=self.breed_var,
            values=self.breeds,
            state="readonly",
            width=18,
        )
        self.breed_combo.pack(side="left")
        self.breed_combo.bind("<<ComboboxSelected>>", self.on_breed_selected)
        self.breed_combo.set(self.breeds[0] if self.breeds else "")

        ttk.Label(row3, text="오늘 결제금액", width=14, anchor="e").pack(side="left", padx=(20, 3))

        self.payment_entry = ttk.Entry(row3, textvariable=self.payment_display, width=14, justify="right")
        self.payment_entry.pack(side="left")
        self.payment_entry.bind("<FocusOut>", self.on_payment_focus_out)

        pay_state_frame = ttk.Frame(row3)
        pay_state_frame.pack(side="left", padx=(10, 0))

        ttk.Radiobutton(
            pay_state_frame,
            text="결제완료",
            variable=self.payment_state,
            value="paid",
        ).pack(side="left")
        ttk.Radiobutton(
            pay_state_frame,
            text="입금 전",
            variable=self.payment_state,
            value="pending",
        ).pack(side="left", padx=(5, 0))

        # 4행: 기타 품종 직접 입력
        row4 = ttk.Frame(info_frame)
        row4.pack(fill="x", pady=2)

        ttk.Label(row4, text="기타 품종 직접 입력", width=14, anchor="e").pack(side="left", padx=(0, 3))
        self.breed_other_entry = ttk.Entry(row4, textvariable=self.breed_other_var, width=30, state="disabled")
        self.breed_other_entry.pack(side="left")

    def _build_notes_area(self, root_frame):
        text_frame = ttk.LabelFrame(root_frame, text="요구사항 / 특이사항 / 애프터 케어")
        text_frame.pack(fill="both", expand=True, padx=5, pady=5)

        # 고객 요구사항
        ttk.Label(text_frame, text="고객 요구사항").pack(anchor="w")
        self.requirements_text = tk.Text(text_frame, height=4)
        self.requirements_text.pack(fill="x", pady=(0, 5))

        # 미용 중 특이사항
        ttk.Label(text_frame, text="미용 중 특이사항").pack(anchor="w")
        self.notes_text = tk.Text(text_frame, height=4)
        self.notes_text.pack(fill="x", pady=(0, 5))

        # 애프터 케어
        ttk.Label(text_frame, text="애프터 케어").pack(anchor="w")
        self.aftercare_text = tk.Text(text_frame, height=4)
        self.aftercare_text.pack(fill="both", expand=True)

    def _build_action_area(self, root_frame):
        btn_frame = ttk.Frame(root_frame)
        btn_frame.pack(fill="x", pady=(8, 0))

        run_btn = ttk.Button(btn_frame, text="실행", command=self.on_run)
        run_btn.pack(side="right")

    def _lock_window_size(self):
        """이미지 삽입 후에도 창 크기가 변하지 않도록 고정."""
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        self.minsize(width, height)
        self.maxsize(width, height)

    # ---------- 플레이스홀더 / 입력 보조 ----------

    def _on_customer_focus_in(self, event):
        if self.customer_entry.get() == "010-0000-0000" and self.customer_entry.cget("foreground") == "gray":
            self.customer_entry.delete(0, "end")
            self.customer_entry.config(foreground="black")

    def _on_customer_focus_out(self, event):
        if not self.customer_entry.get().strip():
            self.customer_entry.insert(0, "010-0000-0000")
            self.customer_entry.config(foreground="gray")

    def on_payment_focus_out(self, event=None):
        raw = self.payment_display.get().strip()
        if not raw:
            return
        digits = "".join(ch for ch in raw if ch.isdigit())
        if not digits:
            messagebox.showerror("입력 오류", "결제금액은 숫자만 입력할 수 있습니다.", parent=self)
            self.payment_display.set("")
            return
        try:
            value = int(digits)
            self.payment_display.set(f"{value:,}")
        except ValueError:
            messagebox.showerror("입력 오류", "결제금액을 해석할 수 없습니다.", parent=self)
            self.payment_display.set("")

    def on_breed_selected(self, event=None):
        selected = self.breed_var.get()
        if selected == "기타(직접입력)":
            self.breed_other_entry.config(state="normal")
            self.breed_other_entry.focus_set()
        else:
            self.breed_other_entry.config(state="disabled")
            self.breed_other_var.set("")

    # ---------- 사진 관련 ----------

    def on_photo_click(self, which: str):
        filetypes = [
            ("이미지 파일", "*.jpg *.jpeg *.png *.JPG *.JPEG *.PNG"),
            ("모든 파일", "*.*"),
        ]
        path = filedialog.askopenfilename(title="사진 선택", filetypes=filetypes)
        if path:
            self.set_photo(which, path)

    def on_photo_drop(self, event, which: str):
        # DND_FILES는 여러 개일 수 있음. 첫 번째만 사용.
        data = event.data
        if " " in data and not (data.startswith("{") and data.endswith("}")):
            # 공백 포함 경로 처리용
            parts = self.split_dnd_files(data)
        else:
            parts = [data.strip("{}")]

        if not parts:
            return
        path = parts[0]
        self.set_photo(which, path)

    @staticmethod
    def split_dnd_files(data: str):
        # {C:\path with space\file.jpg} C:\path2\file2.jpg
        result = []
        current = ""
        in_brace = False
        for ch in data:
            if ch == "{":
                in_brace = True
                current = ""
            elif ch == "}":
                in_brace = False
                result.append(current)
                current = ""
            elif ch == " " and not in_brace:
                if current:
                    result.append(current)
                    current = ""
            else:
                current += ch
        if current:
            result.append(current)
        return result

    def set_photo(self, which: str, path: str):
        if not os.path.isfile(path):
            messagebox.showerror("파일 오류", "파일을 찾을 수 없습니다.", parent=self)
            return

        ext = os.path.splitext(path)[1].lower()
        if ext not in (".jpg", ".jpeg", ".png"):
            messagebox.showerror("파일 오류", "JPG, JPEG, PNG 형식만 사용할 수 있습니다.", parent=self)
            return

        try:
            thumb = load_image_for_thumbnail(path)
        except Exception as e:
            messagebox.showerror("이미지 오류", f"이미지를 불러오는 중 오류가 발생했습니다.\n\n{e}", parent=self)
            return

        if which == "before":
            self.before_path.set(path)
            self.before_thumb = thumb
            self.before_label.config(image=self.before_thumb, text="")
        else:
            self.after_path.set(path)
            self.after_thumb = thumb
            self.after_label.config(image=self.after_thumb, text="")

    # ---------- 실행 ----------

    def on_run(self):
        try:
            # 1. 기본 입력 검증
            before = self.before_path.get()
            after = self.after_path.get()

            if not before or not after:
                messagebox.showerror("입력 오류", "미용 전/후 사진을 모두 선택해주세요.", parent=self)
                return

            dog_name = self.dog_name.get().strip()
            owner_name = self.owner_name.get().strip()
            raw_customer = self.customer_raw.get().strip()
            if raw_customer == "010-0000-0000" and self.customer_entry.cget("foreground") == "gray":
                raw_customer = ""

            if not dog_name or not owner_name or not raw_customer:
                messagebox.showerror("입력 오류", "강아지 이름, 보호자 이름, 고객번호를 모두 입력해주세요.", parent=self)
                return

            # 고객번호: 숫자만 추출, 나머지 문자 있으면 경고
            digits_only = "".join(ch for ch in raw_customer if ch.isdigit())
            if not digits_only:
                messagebox.showerror("입력 오류", "고객번호는 숫자를 포함해야 합니다.", parent=self)
                return
            if any(not (ch.isdigit() or ch in "- ") for ch in raw_customer):
                messagebox.showerror("입력 오류", "고객번호에는 숫자와 '-'만 사용해주세요.", parent=self)
                return
            customer_no = digits_only

            # 품종
            selected_breed = self.breed_var.get().strip()
            if selected_breed == "기타(직접입력)":
                breed = self.breed_other_var.get().strip()
                if not breed:
                    messagebox.showerror("입력 오류", "기타 품종을 직접 입력해주세요.", parent=self)
                    return
            else:
                breed = selected_breed

            # 결제금액
            raw_pay = self.payment_display.get().strip()
            if raw_pay:
                digits = "".join(ch for ch in raw_pay if ch.isdigit())
                if not digits:
                    messagebox.showerror("입력 오류", "결제금액은 숫자만 입력할 수 있습니다.", parent=self)
                    return
                payment_amount = int(digits)
            else:
                payment_amount = 0

            payment_state = "결제완료" if self.payment_state.get() == "paid" else "입금 전"

            style_today = self.style_today.get().strip()
            requirements = self.requirements_text.get("1.0", "end").strip()
            notes = self.notes_text.get("1.0", "end").strip()
            aftercare = self.aftercare_text.get("1.0", "end").strip()

            # 2. 폴더 / 파일명 생성
            now = datetime.now()
            dt_str = now.strftime("%Y%m%d%H%M")  # 202512071434 형식

            folder_name = f"{customer_no} - {owner_name} - {dog_name}"
            folder_name = sanitize_for_path(folder_name)

            base_dir = os.path.join(os.path.abspath(os.path.dirname(__file__)), OUTPUT_ROOT)
            dest_folder = os.path.join(base_dir, folder_name)
            os.makedirs(dest_folder, exist_ok=True)

            # 파일명 공통 prefix
            prefix = f"{customer_no} - {owner_name} - {dog_name} - {dt_str}"
            prefix = sanitize_for_path(prefix)

            # 파일명 길이 체크 (확장자/미용전/후 포함해서 100자 초과 금지)
            ext_before = os.path.splitext(before)[1].lower()
            ext_after = os.path.splitext(after)[1].lower()

            name_before = f"{prefix} - 미용전{ext_before}"
            name_after = f"{prefix} - 미용후{ext_after}"

            if len(name_before) > 100 or len(name_after) > 100:
                messagebox.showerror(
                    "파일명 오류",
                    "생성되는 파일명이 100자를 초과합니다.\n"
                    "강아지 이름이나 보호자 이름, 고객번호를 줄여주세요.",
                    parent=self,
                )
                return

            dest_before = os.path.join(dest_folder, name_before)
            dest_after = os.path.join(dest_folder, name_after)

            # 3. 파일 복사 (PIL로 다시 저장 → 형식 깨끗하게)
            self.save_image_copy(before, dest_before)
            self.save_image_copy(after, dest_after)

            # 4. 엑셀 기록
            record_time = now.strftime("%Y-%m-%d %H:%M")

            row = [
                record_time,
                customer_no,
                owner_name,
                dog_name,
                breed,
                style_today,
                requirements,
                notes,
                aftercare,
                payment_amount,
                payment_state,
                os.path.basename(dest_before),
                os.path.basename(dest_after),
            ]
            append_excel_row(os.path.join(os.path.dirname(__file__), EXCEL_FILE), row)

            # 5. 성공 메시지 & 입력 초기화
            messagebox.showinfo(
                "완료",
                f"사진과 기록이 저장되었습니다.\n\n폴더 위치:\n{dest_folder}",
                parent=self,
            )

            self.reset_inputs()

        except Exception:
            err_text = traceback.format_exc()
            self.show_unexpected_error(err_text)

    def save_image_copy(self, src, dest):
        """원본 이미지를 다시 열어 JPEG/PNG로 깨끗하게 저장."""
        im = Image.open(src)
        ext = os.path.splitext(dest)[1].lower()
        if im.mode in ("RGBA", "LA"):
            bg = Image.new("RGB", im.size, (255, 255, 255))
            alpha = im.split()[-1]
            bg.paste(im, mask=alpha)
            im = bg
        else:
            im = im.convert("RGB")

        if ext in (".jpg", ".jpeg"):
            im.save(dest, format="JPEG", quality=95)
        else:
            im.save(dest, format="PNG")

    def reset_inputs(self):
        """성공 후 입력 초기화."""
        # 사진 포함해서 다 비우기
        self.before_path.set("")
        self.after_path.set("")
        self.before_thumb = None
        self.after_thumb = None
        self.before_label.config(image="", text="미용 전 사진\n(여기로 드롭 또는 클릭)")
        self.after_label.config(image="", text="미용 후 사진\n(여기로 드롭 또는 클릭)")

        self.dog_name.set("")
        self.owner_name.set("")
        self.customer_raw.set("")
        self.customer_entry.delete(0, "end")
        self.customer_entry.insert(0, "010-0000-0000")
        self.customer_entry.config(foreground="gray")

        self.style_today.set("")
        self.breed_var.set(self.breeds[0] if self.breeds else "")
        self.breed_other_var.set("")
        self.breed_other_entry.config(state="disabled")

        self.payment_display.set("")
        self.payment_state.set("paid")

        self.requirements_text.delete("1.0", "end")
        self.notes_text.delete("1.0", "end")
        self.aftercare_text.delete("1.0", "end")

    def show_unexpected_error(self, err_text: str):
        """예상치 못한 오류용 팝업 (복사 버튼 포함)."""
        win = tk.Toplevel(self)
        win.title("예상치 못한 오류")
        win.geometry("520x260")
        win.resizable(False, False)

        frame = ttk.Frame(win, padding=10)
        frame.pack(fill="both", expand=True)

        ttk.Label(
            frame,
            text="작업 중 예기치 못한 오류가 발생했습니다.\n"
            "아래 오류 내용을 개발자에게 전달해 주세요.",
            justify="left",
        ).pack(anchor="w")

        text_box = tk.Text(frame, height=6, wrap="word")
        text_box.pack(fill="both", expand=True, pady=5)
        text_box.insert("1.0", err_text)
        text_box.config(state="disabled")

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(anchor="e", pady=(5, 0))

        def copy_error():
            self.clipboard_clear()
            self.clipboard_append(err_text)
            messagebox.showinfo("복사 완료", "오류 기록이 클립보드에 복사되었습니다.", parent=win)

        ttk.Button(btn_frame, text="오류 기록 복사", command=copy_error).pack(side="right", padx=5)
        ttk.Button(btn_frame, text="닫기", command=win.destroy).pack(side="right")

        win.transient(self)
        win.grab_set()
        self.wait_window(win)


if __name__ == "__main__":
    breeds = load_breeds_or_die(BREEDS_FILE)
    app = DogPhotoTool(breeds)
    app.mainloop()
